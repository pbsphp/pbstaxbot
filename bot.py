# coding: utf-8

import os
import io
import re
import telebot

from openpyxl import Workbook
from openpyxl.styles import Font

from telebot import types

import taxcom_api
import qr_parsing


bot_token = os.getenv('TG_TOKEN')
if bot_token is None:
    raise RuntimeError('TG_TOKEN env variable is not set.')


bot = telebot.TeleBot(bot_token)


# Key - chat id, value - dialog coroutine.
dialogs_map = {}


# Telegram bot keyboard
keyboard_markup = types.ReplyKeyboardMarkup(
    row_width=3, one_time_keyboard=True)
keyboard_markup.row(
    types.KeyboardButton('Мае'),
    types.KeyboardButton('Обои'),
    types.KeyboardButton('Ево'),
)


def make_report(data):
    """Build XLSX file from data with sums and factors.
    """
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 50

    ws.append(['Name', 'Price', 'Factor'])

    pos = 1
    for item in data:
        ws.append([item['name'], item['price'], item['factor']])
        pos += 1

    ws.append([])
    ws.append([])

    ws.append(['Total', f'=SUM(B2:B{pos})'])
    ws.append(['To pay', f'=SUMPRODUCT(B2:B{pos} * C2:C{pos})'])

    for cell in ws['A1'], ws['B1'], ws['C1']:
        cell.font = Font(bold=True)

    report = io.BytesIO()
    wb.save(report)
    report.seek(0)
    report.name = 'tax_report.xlsx'

    return report


def dialog():
    """Create dialog coroutine.
    """

    # Get input data as FP/Sum or image
    while True:
        message = yield

        if message.photo:
            file_size_obj = message.photo[-1]
            file_obj = bot.get_file(file_size_obj.file_id)
            buf = io.BytesIO(bot.download_file(file_obj.file_path))

            qr_data = qr_parsing.parse(buf)
            fp = qr_data.get('fp')
            s = qr_data.get('s')

            if not fp or not s:
                bot.send_message(
                    message.chat.id,
                    'Формат: ФП/Сумма (через слеш) или картинка чека.',
                )
            else:
                break

        elif message.text:
            try:
                fp, s = message.text.split('/', 1)
            except ValueError:
                bot.send_message(
                    message.chat.id,
                    'Формат: ФП/Сумма (через слеш) или картинка чека.',
                )
            else:
                if (not re.match(r'^\d{10}$', fp) or
                        not re.match(r'^\d+(\.\d{1,2})?$', s)):
                    bot.send_message(
                        message.chat.id,
                        'ФП должен быть 10 цифр, а Сумма - числом. Ты втираешь какую-то дичь. Но тебе виднее.'
                    )
                break

        else:
            bot.send_message(
                message.chat.id,
                'Формат: ФП/Сумма (через слеш) или картинка чека.',
            )

    # Try to get results
    try:
        data = taxcom_api.get(fp, s)
    except Exception as exc:
        bot.send_message(message.chat.id, f'Все наебнулось. Сори. {exc}')
    else:
        # Iterate over results and get factors
        for idx, item in enumerate(data, start=1):
            while True:
                bot.send_message(
                    message.chat.id,
                    f'{idx}. {item["name"]} ({item["price"]} р.)',
                    reply_markup=keyboard_markup,
                )

                message = yield

                factors_map = {
                    'Мае': 0.0,
                    'Обои': 0.5,
                    'Ево': 1.0,
                }

                try:
                    item['factor'] = factors_map[message.text]
                except KeyError:
                    bot.send_message(message.chat.id, 'Ааа?')
                else:
                    break

        # Print results
        report = make_report(data)
        total_sum = sum(x['price'] for x in data)
        invoice = sum(x['price'] * x['factor'] for x in data)

        try:
            bot.send_message(
                message.chat.id,
                f'Итого: {total_sum}, отжать: {invoice}.'
            )
        except Exception as exc:
            bot.send_message(
                message.chat.id,
                f'Сборка отчета наебнулась: {exc}',
            )
        else:
            bot.send_document(message.chat.id, report)


@bot.message_handler(
    func=lambda message: True, content_types=['text', 'photo'])
def send_welcome(message):
    try:
        dialogs_map[message.chat.id]
    except KeyError:
        dialogs_map[message.chat.id] = dialog()
        dialogs_map[message.chat.id].send(None)

    try:
        dialogs_map[message.chat.id].send(message)
    except StopIteration:
        del dialogs_map[message.chat.id]


bot.polling()
