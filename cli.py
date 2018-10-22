# coding: utf-8

from openpyxl import Workbook
from openpyxl.styles import Font

from taxcom_api.taxcom_api import get


if __name__ == '__main__':
    import re

    fp = input('FP: ')
    s = input('S: ')

    if not re.match(r'^\d{10}$', fp):
        print(f'Warning: {fp} does not match \\d{{10}}.')

    if not re.match(r'^\d+(\.\d{1,2})?$', s):
        print(f'Warning: {s} does not match \\d+(\\.\\d{{1,2}})?.')

    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 50

    ws.append(['Name', 'Price', 'Factor'])

    pos = 1
    for item in get(fp, s):
        ws.append([item['name'], item['price'], 0.5])
        pos += 1

    ws.append([])
    ws.append([])

    ws.append(['Total', f'=SUM(B2:B{pos})'])
    ws.append(['To pay', f'=SUMPRODUCT(B2:B{pos} * C2:C{pos})'])

    for cell in ws['A1'], ws['B1'], ws['C1']:
        cell.font = Font(bold=True)

    wb.save('/tmp/test.xlsx')
